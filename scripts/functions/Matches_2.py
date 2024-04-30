import json
import pandas as pd
from openpyxl.styles import Alignment

# Load upcoming matches data
with open(r'C:\Users\soren\Projects\ValAlgo\data\JSON\upcoming_matches.json', 'r') as file:
    upcoming_matches = json.load(file)

# Load stats data
with open(r'C:\Users\soren\Projects\ValAlgo\data\JSON\stats.json', 'r') as file:
    stats_data = json.load(file)

# Parse the upcoming matches to organize by match ID
match_details = {}
for match in upcoming_matches:
    match_details[match['id']] = [team['name'] for team in match['teams']]

# Parse stats data to organize by team name
team_stats = {}
for team in stats_data:
    team_name = team['info']['name']
    team_stats[team_name] = [{'user': player['user'], 'stats': {key: value for key, value in player.items() if 'stats' in key}} for player in team['players']]

# Define map order and statistics of interest
map_order = ["ascent", "bind", "breeze", "icebox", "lotus", "split", "sunset"]
stats_of_interest = ['KPR', 'RND']  # Stats you want to include

# Directory for saving Excel files
output_directory = r'C:\Users\soren\Projects\ValAlgo\maps_bans_picks\Excel'

# Create Excel file for each match ID
for match_id, teams in match_details.items():
    if all(team in team_stats for team in teams):
        # Prepare data for DataFrame
        data = []
        for team in teams:
            for player in team_stats[team]:
                player_data = {'Team': team, 'Username': player['user']}
                for map_name in map_order:
                    stats_key = f"{map_name}_stats"
                    for stat in stats_of_interest:
                        full_stat_key = f"{map_name}_{stat}"
                        if stats_key in player['stats'] and stat in player['stats'][stats_key]:
                            player_data[full_stat_key] = player['stats'][stats_key][stat]
                        else:
                            player_data[full_stat_key] = None  # Use None for missing stats
                data.append(player_data)

        # Convert to DataFrame
        df = pd.DataFrame(data)

        # Generate filename from team names and match ID
        team_names = ' vs '.join(teams)
        filename = f"{match_id} - {team_names}.xlsx"
        file_path = f'{output_directory}\\{filename}'

        # Save to Excel and set column width
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            sheet_name = 'Match Stats'
            # Write DataFrame to Excel starting from row 3 (index 2 in zero-indexing) to leave room for headers
            df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False, header=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Set column headers manually in row 2
            header_row = 2
            headers = ['Team', 'Username'] + [stat for map_name in map_order for stat in stats_of_interest]
            for col_num, header in enumerate(headers, start=1):
                worksheet.cell(row=header_row, column=col_num).value = header

            # Merge cells for map names in row 1
            current_column = 3  # Assuming 'Team' and 'Username' are in columns 1 and 2
            for map_name in map_order:
                start_column = current_column
                end_column = start_column + len(stats_of_interest) - 1
                worksheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
                worksheet.cell(row=1, column=start_column).value = map_name
                current_column = end_column + 1

            # Set column widths
            for col in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[chr(64 + col)].width = 15

            # Set alignment for all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')




