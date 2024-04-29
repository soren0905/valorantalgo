import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

def load_data(json_path):
    with open(json_path, 'r', encoding='utf-8') as file:
        return json.load(file)

def create_dataframe(teams):
    rows = []
    for team in teams:
        team_name = team['info']['name']
        for player in team['players']:
            row = {'Team': team_name, 'Player': player['user']}
            for map_stat in player.keys():
                if map_stat not in ['id', 'user']:
                    for stat in ['KPR', 'RND']:
                        key = f'{map_stat}_{stat}'
                        if player[map_stat] and stat in player[map_stat]:
                            row[key] = player[map_stat][stat]
                        else:
                            row[key] = None
                            print(f"Missing or zero data for {player['user']} on {map_stat} for {stat}")
            rows.append(row)
    return pd.DataFrame(rows)


def save_to_excel(dataframe, filename):
    # Ensure the directory exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Set the starting row for the DataFrame
    start_row = 4

    # Convert DataFrame to rows in Excel starting from the fourth row
    for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Apply styles for header row
            if r_idx == start_row:
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                cell.font = Font(bold=True, color="000000")
                cell.alignment = Alignment(horizontal="center")
            # Set width of the column
            column_letter = cell.column_letter
            ws.column_dimensions[column_letter].width = 20
            # Set border for each cell
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border

    # Apply simplified headers
    for col in ws.iter_cols(min_row=start_row, max_row=start_row, min_col=3, max_col=ws.max_column):
        for cell in col:
            # Split the cell value and take only the stat part which should be 'RND' or 'KPR'
            parts = cell.value.split('_')
            if len(parts) > 1 and parts[-1] in ['RND', 'KPR']:
                cell.value = parts[-1]

    # Set the title cells for the map names
    map_ranges = ['C3:D3', 'E3:F3', 'G3:H3', 'I3:J3', 'K3:L3', 'M3:N3', 'O3:P3']
    map_names = ['Ascent', 'Bind', 'Icebox', 'Lotus', 'Split', 'Breeze', 'Sunset']
    for map_range, map_name in zip(map_ranges, map_names):
        start_column = ord(map_range.split(':')[0][0]) - 64
        ws.merge_cells(start_row=start_row-1, start_column=start_column, end_row=start_row-1, end_column=start_column+1)
        cell = ws.cell(row=start_row-1, column=start_column, value=map_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Save the Excel file
    wb.save(filename)
    print(f"Data saved to Excel file at: {filename}")

def main():
    project_root = r'C:\Users\soren\Projects\ValAlgo'
    json_path = os.path.join(project_root, 'data', 'JSON', 'upcoming_with_stats.json')

    data = load_data(json_path)
    print(data)  # Add this to check how data looks immediately after loading

    # Split data into pairs of teams
    pairs = [data[i:i+2] for i in range(0, len(data), 2)]

    for pair in pairs:
        team_names = '_and_'.join(team['info']['name'].replace(' ', '_') for team in pair)
        dataframe = create_dataframe(pair)
        excel_path = os.path.join(project_root, 'data', 'Excel', f'{team_names}_stats.xlsx')
        save_to_excel(dataframe, excel_path)

if __name__ == '__main__':
    main()

